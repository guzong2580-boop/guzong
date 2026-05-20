# -*- coding: utf-8 -*-
"""
교육생 목록.xlsx  →  교육생구글시트용.xlsx → Google Sheets 자동 동기화

[규칙]
1. 원본 '교육생 목록.xlsx'의 헤더 행을 찾는다 (이름·개강반 컬럼 있는 행)
2. 헤더 + 데이터 전체 컬럼을 그대로 시트에 푸시 (관리자 상세 페이지용)
3. 추가로 첫 두 컬럼은 [이름, 개강반(가장 최근 개강반 단일 값)]으로 강제 정렬
   → 기존 App.js의 students 파싱(cols[0]=name, cols[1]=개강반)과 호환
4. 이름이 비어있는 행은 건너뛴다
5. service-account.json 이 있으면 Google Sheets 에도 자동 push

[출력 컬럼 순서]
이름 | 개강반(최근) | 설명 | 개강반1 | 개강반2 | 개강반3 | 실습 | 비용 | 비고 | 실습2 | ...

[사용법]
$ python sync.py
"""
import openpyxl
import os, sys, shutil
from datetime import datetime

SHEET_ID = '1n28IPdaxcc4ty5C7E_DiPTVb5RTAzLQddyMlzYHeJ74'
WORKSHEET_GID = 1230895908

# 개인정보 보호 — 공개 CSV에 올리지 않을 컬럼
SENSITIVE_COLS = {'설명', '실습', '비용'}

# 콘솔 한글 출력
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, '교육생 목록.xlsx')
DST = os.path.join(BASE, '교육생구글시트용.xlsx')


def find_header(ws, max_search_rows=10, max_cols=20):
    """헤더 행 위치와 모든 컬럼 위치를 찾는다"""
    for r in range(1, max_search_rows + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)]
        if '이름' in row and '개강반' in row:
            # 헤더 셀들 (None 제거하지 않고 위치 유지, 민감 컬럼 제외)
            header_cells = []
            for i, v in enumerate(row):
                if v is not None and str(v).strip():
                    name = str(v).strip()
                    if name in SENSITIVE_COLS:
                        continue
                    header_cells.append((i + 1, name))
            # 이름·개강반 컬럼 인덱스
            name_col = next(c for c, v in header_cells if v == '이름')
            gan_cols = [c for c, v in header_cells if v.startswith('개강반')]
            return r, header_cells, name_col, gan_cols
    return None, None, None, None


def dedupe_headers(headers):
    """중복 헤더에 숫자 접미사. 이미 사용된 이름과 충돌 시 다음 숫자 사용
    (실습, 실습 → 실습, 실습2)
    (개강반, 개강반2 가 들어와도 두 번째 개강반은 '개강반3'으로)
    """
    out = []
    used = set()
    counter = {}
    for h in headers:
        if h not in used:
            out.append(h)
            used.add(h)
            counter[h] = 1
        else:
            # 충돌하지 않는 숫자 찾기
            base = h
            n = counter.get(base, 1) + 1
            candidate = f'{base}{n}'
            while candidate in used:
                n += 1
                candidate = f'{base}{n}'
            out.append(candidate)
            used.add(candidate)
            counter[base] = n
    return out


def main():
    if not os.path.exists(SRC):
        print(f'X 원본 파일이 없습니다: {SRC}')
        sys.exit(1)

    print('=' * 60)
    print('교육생 명단 동기화 (전체 컬럼)')
    print('=' * 60)
    print(f'원본:  {SRC}')
    print(f'대상:  {DST}')
    print()

    # 원본 읽기
    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    ws_src = wb_src.active

    header_row, header_cells, name_col, gan_cols = find_header(ws_src)
    if header_row is None:
        print('X 헤더 행(이름·개강반)을 찾지 못했습니다.')
        sys.exit(1)

    # 원본 헤더 순서 보존 (None 제외) — (col_index, header_name) 리스트
    src_headers = [h for _, h in header_cells]
    src_cols = [c for c, _ in header_cells]
    print(f'  헤더 행: {header_row}')
    print(f'  컬럼 ({len(src_headers)}개): {src_headers}')
    print(f'  이름 열: {name_col} / 개강반 열들: {gan_cols}')
    print()

    # 기존 결과 백업
    if os.path.exists(DST):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        bak = DST + f'.{ts}.bak'
        shutil.copy2(DST, bak)
        print(f'  백업: {os.path.basename(bak)}')
        print()

    # 데이터 추출
    students = []
    for r in range(header_row + 1, ws_src.max_row + 1):
        name = ws_src.cell(row=r, column=name_col).value
        if not name or not str(name).strip():
            continue

        # 가장 최근 개강반
        selected_gan = None
        for c in reversed(gan_cols):
            val = ws_src.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                selected_gan = str(val).strip()
                break

        if not selected_gan:
            print(f'  ! {name}: 개강반 정보 없음 - 건너뜀')
            continue

        # 전체 컬럼 값 (원본 헤더 순서대로)
        full_row = []
        for c in src_cols:
            v = ws_src.cell(row=r, column=c).value
            if v is None:
                full_row.append('')
            else:
                # datetime을 문자열로
                if hasattr(v, 'strftime'):
                    full_row.append(v.strftime('%Y.%m.%d'))
                else:
                    full_row.append(str(v).strip())

        students.append({
            'name': str(name).strip(),
            'gan': selected_gan,
            'full': full_row,
        })

    # 출력 헤더 정의: [이름, 개강반(최근)] + 원본 컬럼 (이름 제외, 충돌 방지 리네임)
    # → App.js에서 cols[0]=name, cols[1]=개강반 으로 파싱하던 기존 로직 호환
    # 원본 '개강반/개강반2/개강반3'은 차수 컬럼으로 의미 구분 → '개강반차수1/2/3'
    rename_map = {
        '개강반': '개강반차수1',
        '개강반2': '개강반차수2',
        '개강반3': '개강반차수3',
    }
    other_headers_raw = [rename_map.get(h, h) for h in src_headers if h != '이름']
    other_headers = dedupe_headers(other_headers_raw)  # 남은 중복 (예: 실습 2개) 처리
    final_headers = ['이름', '개강반'] + other_headers

    # 결과 미리보기
    print(f'=== 추출 결과 ({len(students)}명) ===')
    print('  헤더:', final_headers)
    for i, st in enumerate(students[:5], 1):
        print(f'  {i:2}. {st["name"]:8} | {st["gan"]} | {st["full"]}')
    if len(students) > 5:
        print(f'  ... (총 {len(students)}명)')
    print()

    # 출력 데이터 행 빌드
    out_rows = []
    for st in students:
        # name 컬럼 제외한 원본 값들
        others = [v for c, v in zip(src_cols, st['full']) if c != name_col]
        out_rows.append([st['name'], st['gan']] + others)

    # 구글시트용 파일 작성
    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = 'Sheet1'
    for j, h in enumerate(final_headers):
        ws_dst.cell(row=1, column=j + 1, value=h)
    for i, row in enumerate(out_rows):
        for j, v in enumerate(row):
            ws_dst.cell(row=i + 2, column=j + 1, value=v)
    wb_dst.save(DST)

    print('OK 로컬 동기화 완료')
    print()

    # Google Sheets 푸시
    key_file = os.path.join(BASE, 'service-account.json')
    if not os.path.exists(key_file):
        print('i  service-account.json 없음 — Google Sheets push 건너뜀')
        return

    print('Google Sheets 푸시 중...')
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = Credentials.from_service_account_file(key_file, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.get_worksheet_by_id(WORKSHEET_GID)

        values = [final_headers] + out_rows
        ws.clear()
        ws.update(values=values, range_name='A1')

        print(f'OK Google Sheets 푸시 완료 ({len(students)}명, {len(final_headers)}컬럼)')
        print('   ~5분 후 https://guzong.vercel.app 에 반영됨')
    except gspread.exceptions.APIError as e:
        print(f'X Google Sheets API 오류: {e}')
        print('   -> 시트가 서비스 계정에 편집자 공유 됐는지 확인:')
        print('   seoul-edu-sync@quantum-toolbox-488410-n4.iam.gserviceaccount.com')
    except Exception as e:
        print(f'X 푸시 실패: {type(e).__name__}: {e}')


if __name__ == '__main__':
    main()
